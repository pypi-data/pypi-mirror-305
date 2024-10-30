import logging
import re
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from ons_metadata_validation.io.cell_template import (
    MetadataCell,
    MetadataValues,
)
from ons_metadata_validation.utils.logger import (
    compress_logging_value,
)

logger = logging.getLogger()


def apply_version_changes(
    template_map: Dict, delta_table: Dict, target_version: str
) -> Dict:
    """_summary_

    Args:
        template_map (Dict): _description_
        delta_table (Dict): _description_
        target_version (str): _description_

    Returns:
        Dict: _description_
    """
    for version in delta_table.keys():
        if float(version.strip("v")) >= float(target_version.strip("v")):
            continue
        for key, changes in delta_table[version].items():
            if key in template_map:
                for cell_attr, attr_value in changes.items():
                    template_map[key][cell_attr] = attr_value
            else:
                template_map[key] = changes
    return template_map


def get_matching_cell(md_refs: dict, tab: str, name: str) -> MetadataCell:
    """find the MetadataCell object that houses the information for the tab and
    column name

    Args:
        md_refs (list): List of MetadataCells objects housing the metadata for
            the xlsx file
        tab (str): the tab to search for
        name (str): the column to search for

    Raises:
        ValueError: there should only be one location per MetadataCell, so if
            returns more or less than one raises error

    Returns:
        MetadataCell: the metadata cell that matches the tab and name
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    key = _get_template_dict_key(tab, name)
    if key in md_refs:
        return md_refs[key]
    raise KeyError(f"No entry that matches key: {key}")


def check_metadata_vs_map(
    filepath: str, md_ref: Dict[str, MetadataCell]
) -> Dict[str, str]:
    """sanity check function that ensures the xlsx workbook matches the
    md_ref provided. E.g. the "File format" cell is on tab "Dataset File"
    in cell "E7"

    Args:
        filepath (str): the filepath to the workbook
        md_ref (List[MetadataCell]): the template map as a list of metadata cells

    Raises:
        ValueError: raises exceptions and prints the error as it runs

    Returns:
        Dict[str, str]: dict of the values that don't match
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    wb = load_workbook(filename=filepath)
    matches: Dict[str, List] = {}

    for item in md_ref.values():
        try:
            template_value = wb[item.tab][item.ref_cell].value
            values_match = template_value == item.name
            matches[item.name] = [template_value, values_match]

            if not values_match:
                raise ValueError(
                    f"Values in map and metadata doesn't match for {item}"
                )
        except ValueError as e:
            logger.error(e)

    non_matches: Dict[str, str] = {
        k: v[0] for k, v in matches.items() if v[1] is False
    }
    return non_matches


def load_metadata_cells(template_map: Dict) -> Dict[str, MetadataCell]:
    """load the template_map to MetadataCell objects, these perform basic
    validation on init to catch any issues with the template that may have
    happened in storage

    Args:
        template_map (List[Dict]): list of dictionaries that map the metadata
            template

    Returns:
        (List[MetadataCell]): list of metadata cell objects
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    values = {}
    for key, data in template_map.items():
        values[key] = MetadataCell(**data)
        logger.debug(f"added {data['tab']} {data['name']}")
    return values


def fill_metadata_values(
    wb: Workbook, metadata_values: MetadataValues
) -> None:
    """populate the workbook with the metadata values

    Args:
        wb (Workbook): the workbook object to populate, this is done inplace
        metadata_values (MetadataValues): the object to populate
    """
    logger.debug(f"metadata_values = {metadata_values}")
    for idx, value in enumerate(metadata_values.values):
        value_col = metadata_values.cell.value_col
        value_idx = metadata_values.cell.row_start + idx
        wb[metadata_values.cell.tab][f"{value_col}{value_idx}"] = value


def extract_all_metadata_values(
    wb: Workbook, md_refs: Dict[str, MetadataCell]
) -> Dict[str, MetadataValues]:
    """wrapper function for _extract_metadata_values(). Applies that function
    to all MetadataCells in List.

    Args:
        wb (Workbook): the workbook object to extract data from
        md_refs (List[MetadataCell]): the list of MetadataCells with the
            references to extract

    Returns:
        List[MetadataValues]: _description_
    """
    logger.debug(f"md_refs = {md_refs}")
    metadata_values = {}

    for key, cell in tqdm(md_refs.items(), "extracting metadata values"):
        metadata_values[key] = _extract_metadata_values(wb, cell)

    return metadata_values


def _extract_metadata_values(
    wb: Workbook, cell: MetadataCell
) -> MetadataValues:
    """extract the values from a metadata cell

    Args:
        wb (Workbook): the metadata workbook
        cell (MetadataCell): the metadata cell to extract values from

    Returns:
        MetadataValues: the metadata value class with the values populated
    """
    logger.debug(f"cell = {cell}")
    n_rows = wb[cell.tab].max_row - cell.row_start

    values = []

    if cell.column:
        for i in range(n_rows):
            values.append(
                wb[cell.tab][f"{cell.value_col}{cell.row_start+i}"].value
            )
    else:
        values.append(wb[cell.tab][f"{cell.value_col}{cell.row_start}"].value)
    return MetadataValues(cell, values)


def _get_template_dict_key(inp_tab: str, inp_name: str) -> str:
    """_summary_

    Args:
        inp_tab (str): _description_
        inp_name (str): _description_

    Returns:
        str: _description_
    """
    tab = _remove_special_chars(inp_tab)
    name = (
        _remove_special_chars(inp_name, "_")
        .replace(" ", "_")
        .strip("_")
        .lower()
    )
    return f"{tab}_{name}"


def _remove_special_chars(inp_string: str, replace: str = "") -> str:
    """_summary_

    Args:
        inp_string (str): _description_
        replace (str, optional): _description_. Defaults to "".

    Returns:
        str: _description_
    """
    return re.sub(r"[^0-9a-zA-Z_]+", replace, inp_string)
