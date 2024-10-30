import itertools
import logging
from typing import Dict, List, Tuple

import attrs
import pandas as pd
from attrs.validators import instance_of
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from ons_metadata_validation.fixing.fixing_utils import fix_main
from ons_metadata_validation.io.cell_template import Fail, MetadataValues
from ons_metadata_validation.io.input_functions import (
    apply_version_changes,
    check_metadata_vs_map,
    extract_all_metadata_values,
    load_metadata_cells,
)
from ons_metadata_validation.processing.developing import (
    md_values_to_df,
    search_md_values,
)
from ons_metadata_validation.reference.delta_table import DELTA_TABLE
from ons_metadata_validation.reference.role_based_configs import (
    ROLE_CONFIGS,
)
from ons_metadata_validation.reference.v2_template import V2_TEMPLATE
from ons_metadata_validation.utils.logger import (
    compress_logging_value,
)
from ons_metadata_validation.validation import (
    comparative_validations as comp_val,
)

logger = logging.getLogger()


@attrs.define
class MetadataProcessor:
    md_filepath: str = attrs.field(validator=[instance_of(str)])
    variable_check_set: str = attrs.field(validator=[instance_of(str)])
    save_report: bool = attrs.field(validator=[instance_of(bool)])
    save_corrected_copy: bool = attrs.field(validator=[instance_of(bool)])
    wb: Workbook = attrs.field(validator=[instance_of(Workbook)], init=False)
    target_version: str = attrs.field(validator=[instance_of(str)], init=False)
    template_map: dict = attrs.field(validator=[instance_of(dict)], init=False)
    metadata_values: dict = attrs.field(
        validator=[instance_of(dict)], init=False
    )
    tab_lens: dict = attrs.field(validator=[instance_of(dict)], init=False)
    fails_df: pd.DataFrame = attrs.field(
        validator=[instance_of(pd.DataFrame)], init=False
    )

    def main_process(self) -> None:
        """the main process for the validation

        Args:
            md_filepath (str): the filepath to the metadata template
            save_report (bool, optional): Whether to save the report. Defaults to True.
            save_corrected_copy (bool, optional): Whether to save a corrected version of the template. Defaults to False.

        Raises:
            ValueError: if metadata map and input template don't match
            RuntimeError: if fails to remove all the null rows
            RuntimeError: if fails to validate all metadata values
            RuntimeError: if fails to save to excel

        Returns:
            Dict[str, pd.DataFrame]: the fails dict of hard and soft fails
        """
        for key, val in locals().items():
            logger.debug(f"{key} = {compress_logging_value(val)}")

        self.wb = load_workbook(self.md_filepath, data_only=True)
        self.target_version = self.wb["Dataset Resource"]["A1"].value
        self.template_map = apply_version_changes(
            V2_TEMPLATE, DELTA_TABLE, self.target_version
        )
        md_refs = load_metadata_cells(self.template_map)

        role_config = ROLE_CONFIGS[self.variable_check_set.lower()]

        non_matches: Dict[str, str] = check_metadata_vs_map(
            self.md_filepath, md_refs
        )
        if non_matches:
            msg = (
                f"metadata file does not match map {non_matches}. "
                "Check if using the correct metadata map or make new one"
            )
            logger.error(msg)
            raise ValueError(msg)

        self.metadata_values = extract_all_metadata_values(self.wb, md_refs)

        if not _remove_null_rows_for_tabs(
            self.metadata_values,
            [
                "Dataset File",
                "Dataset Series",
                "Variables",
                "Codes and Values",
            ],
        ):
            msg = "_remove_null_rows_for_tabs failed to remove all null rows"
            logger.error(msg)
            raise RuntimeError(msg)

        self.tab_lens = get_tab_lens(self.metadata_values)

        if not _validate_metadata_values(self.metadata_values, role_config):
            msg = "_validate_metadata_values failed to validate all metadata values"
            logger.error(msg)
            raise RuntimeError(msg)

        self.fails_df = populate_fail_df(
            self.wb, self.metadata_values, self.template_map
        )
        if self.save_corrected_copy:
            fix_main(self.wb, self.fails_df, self.md_filepath)


def get_tab_lens(metadata_values: Dict[str, MetadataValues]) -> Dict[str, int]:
    """gets the number of rows for each of the main tabs

    Args:
        metadata_values (Dict[str, MetadataValues]): the metadata_values to get the length from

    Returns:
        Dict[str, int]: dict of the tab name and the row counts
    """
    tabs = [
        "Dataset Resource",
        "Dataset File",
        "Dataset Series",
        "Variables",
        "Codes and Values",
    ]

    df_dict = {tab: md_values_to_df(metadata_values, tab) for tab in tabs}
    return {tab: len(df_dict[tab]) for tab in tabs}


def _remove_null_rows_for_tabs(
    metadata_values: Dict[str, MetadataValues], tabs: List[str]
) -> bool:
    """remove rows for all tabs that are all(null or have a formula in them).
    Wrapper function for _remove_null_rows()

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        tabs (List[str]): list of tabs to remove nulls from

    Returns:
        bool: True if all are successful
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    successes = []

    for tab in tabs:
        successes.append(_remove_null_rows(metadata_values, tab))

    return all(successes)


def _remove_null_rows(
    metadata_values: Dict[str, MetadataValues], tab: str
) -> bool:
    """remove rows for given tab that are all(null or have a formula in them).

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        tab (str): tab to remove nulls from

    Returns:
        bool: True if successful
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    try:
        tab_mvs = _get_md_values_by_tab(metadata_values, tab)
        tab_values = [mv.values for mv in tab_mvs]
        tab_records = list(zip(*tab_values))

        null_idx = []

        for i, record in tqdm(
            enumerate(tab_records), f"finding null rows for {tab} tab"
        ):
            # all the Nones are cast to str, and formulas are dragged down to a certain
            # row, so if all None or a formula we can be roughly sure it's a null row
            if all(
                [
                    str(value) in ["None", " ", ""]
                    or str(value).startswith("=")
                    for value in record
                ]
            ):
                null_idx.append(i)

        for mv in tqdm(tab_mvs, f"removing null rows for {tab} tab"):
            # reversed to avoid deleting index 1 and then the index 2 becoming index 1 etc
            for i in sorted(null_idx, reverse=True):
                logger.debug(
                    f"deleting value | idx: {i} | tab: {mv.cell.tab} | name: {mv.cell.name} | value: {mv.values[i]} "
                )
                del mv.values[i]

        return True
    except Exception as e:
        logger.error(f"_remove_null_rows() unsuccessful for {tab}: {e}")
        return False


def _get_md_values_by_tab(
    metadata_values: Dict[str, MetadataValues], tab: str
) -> List[MetadataValues]:
    """gets all the MetadataValues objects from a tab

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        tab (str): the tab to get the values from, pulled from the tab attribute of the
            MetadataCell within the MetadataValues object

    Returns:
        List[MetadataValues]: the subset of MetadataValues from that tab
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    return [v for v in metadata_values.values() if v.cell.tab == tab]


def _validate_metadata_values(
    metadata_values: Dict[str, MetadataValues], role_config: List[Tuple[str]]
) -> bool:
    """calls the validate method for each MetadataValues object

    Args:
        metadata_values (Dict[str, MetadataValues]): the dict of extracted metadata_values

    Returns:
        bool: when all are validated
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    for mv in tqdm(metadata_values.values(), "validating cells"):
        if (mv.cell.tab, mv.cell.name) in role_config:
            mv.validate()
    return True


def populate_fail_df(
    wb: Workbook,
    metadata_values: Dict[str, MetadataValues],
    template_map: Dict,
) -> pd.DataFrame:
    """creates a dataframe of each of the fails that

    Args:
        wb (Workbook): the metadata template
        metadata_values (Dict[str, MetadataValues]): the dict of extracted metadata_values
        template_map (Dict): the template map for the template version

    Returns:
        pd.DataFrame: dataframe consisting of fail object attributes
    """
    all_fails = _get_all_fails(metadata_values, template_map)
    populated_fails = list(
        itertools.chain.from_iterable(
            [
                _populate_fail_cell_refs(wb, metadata_values, f)
                for f in tqdm(all_fails, "populating fails")
            ]
        )
    )
    return pd.DataFrame(
        [attrs.asdict(f) for f in populated_fails]
    ).drop_duplicates()


def _get_all_fails(
    metadata_values: Dict[str, MetadataValues], template_map: Dict
) -> List[Fail]:
    """get the fails for all of the metadata values and comparative fails

    Args:
        metadata_values (Dict[str, MetadataValues]): the dict of extracted metadata_values
        template_map (Dict): the template map for the template version

    Returns:
        List[Fail]: list of fail objects for each fail per value
    """
    all_fails = list(
        itertools.chain.from_iterable(
            [mv.hard_fails + mv.soft_fails for mv in metadata_values.values()]
        )
    )
    comp_fails = _get_comparative_fails(metadata_values, template_map)
    all_fails += comp_fails
    return all_fails


def _populate_fail_cell_refs(
    wb: Workbook, metadata_values: Dict[str, MetadataValues], inp_fail: Fail
) -> List[Fail]:
    """get all the fail references and create a single Fail object for each fail and cell reference

    Args:
        wb (Workbook): the metadata template
        metadata_values (Dict[str, MetadataValues]): the dict of extracted metadata_values
        inp_fail (Fail): the individual Fail object to get references for

    Returns:
        List[Fail]: list of fails one for each validation check, value, cell reference
    """
    fails_with_refs = []
    refs = _get_fail_references(wb, metadata_values, inp_fail)
    if refs:
        for ref in refs:
            fail = Fail(**attrs.asdict(inp_fail))
            fail.cell_ref = ref
            fails_with_refs.append(fail)
    # without this else we lose comparative fails that don't have a reference
    else:
        fails_with_refs.append(inp_fail)
    return fails_with_refs


def _get_comparative_fails(
    metadata_values: Dict[str, MetadataValues], template_map: Dict
) -> List[Fail]:
    """wrapper to run all comparative validation checks

    Args:
        metadata_values (List[MetadataValues]): populated MetadataValues objects

    Returns:
        List[Dict]: list of errors in the same format as hard and soft fails
    """
    tabs = [
        "Dataset Resource",
        "Dataset File",
        "Dataset Series",
        "Variables",
        "Codes and Values",
    ]

    df_dict = {tab: md_values_to_df(metadata_values, tab) for tab in tabs}
    # this is the only place that uses pandas, it should be safe from pd version differences
    # these require dataframes as their interface which why they can be run
    # in a loop
    df_comparative_validations = {
        "Dataset File": [
            comp_val.check_DatasetFile_csv_string_identifier,
            comp_val.check_DatasetFile_multiple_files_to_append,
            comp_val.check_DatasetFile_csv_column_separators,
            comp_val.check_DatasetFile_csv_number_of_header_rows,
        ],
        "Variables": [
            comp_val.check_Variables_duplicate_variable_names,
            comp_val.check_Variables_variable_length_precision,
            comp_val.check_Variables_null_values_denoted_by,
            comp_val.check_Variables_foreign_key_file_name,
        ],
        "Codes and Values": [
            comp_val.check_CodesAndValues_duplicate_key_names,
        ],
    }

    mv_comp_vals = [
        comp_val.check_table_names_appear_in_main_tabs,
        comp_val.check_all_unique_values_cols,
        comp_val.check_Variables_CodesAndValues_is_this_a_code,
        comp_val.check_DatasetSeries_length_for_DatasetResource_vs_bigquery_table,
    ]

    comp_errors = []

    for comp_func in mv_comp_vals:
        comp_errors.extend(comp_func(metadata_values, template_map))

    for tab_name, val_funcs in df_comparative_validations.items():
        for val_func in val_funcs:
            comp_errors.extend(val_func(df_dict[tab_name], template_map))

    return comp_errors


def _get_fail_references(
    wb: Workbook, metadata_values: Dict[str, MetadataValues], fail: Fail
) -> List:
    """_summary_

    Args:
        wb (Workbook): the metadata template workbook (not modified)
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        fail (Dict): single fail instance

    Returns:
        str: a joined str of the list of cell references
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    mv = search_md_values(metadata_values, fail.tab, fail.name)

    # is len(mv.values) cleaner? avoids going over loads of empty rows
    # more risky but less cell refs for "None"s that go to the bottom of the sheet
    n_rows = len(mv.values)  # wb[mv.cell.tab].max_row - mv.cell.row_start
    values = []

    if mv.cell.column:
        for n in range(n_rows):
            cell_loc = f"{mv.cell.value_col}{mv.cell.row_start+n}"
            cell_value = wb[mv.cell.tab][cell_loc].value

            # convert the cell_value to str for easy comparison
            if str(cell_value) == str(fail.value):
                values.append(cell_loc)

    else:
        values.append(f"{mv.cell.value_col}{mv.cell.row_start}")

    logger.info(f"fail: {fail} | cell_locs: {values}")
    return values
