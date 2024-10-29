import inspect

# TODO: logging
import os
import re
from datetime import datetime
from types import ModuleType
from typing import Dict, List, Tuple

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook

import ons_metadata_validation.validation._validation_checks as vc

#######################
"""Main function"""
#######################


def outputs_main(processor, save_folder=None, save_commented_copy=True):
    """Write output report and possibly saves a commented copy, too.

    Notes:

    save_commented_copy is currently always set to True, since the interface is still
    under development.
    """

    if save_folder is None:  # by default, save to same folder as input file
        save_folder = os.path.dirname(processor.md_filepath)

    report_wb = write_output_report(processor, save_folder)

    if save_commented_copy:
        comment_wb = create_copy_with_comments(processor, save_folder)

    return report_wb


#######################
"""Support functions"""
#######################


def _strip_col_letter(cell_list: List[str]) -> List[int]:
    """Takes a list of excel cells and removes their shared column reference letter.

    Args:
        cell_list (list[str]): A list of excel cell references, e.g. ['A6','A7','A8']. All cells must belong to the same column.

    Returns:
        num_list (list[int]): A list of the row references of the input cells as ints, e.g. [6,7,8].
    """
    assert isinstance(cell_list, list)
    assert all([isinstance(val, str) for val in cell_list])

    col_letter = "".join(
        [char for char in str(cell_list[0]) if char.isalpha()]
    )
    letter_len = len(col_letter)  # could be 2, e.g. AA
    assert all(
        [val[:letter_len] == col_letter for val in cell_list]
    ), "These cells aren't all from the same column!"
    num_list = [int(val[letter_len:]) for val in cell_list]

    return num_list


def drop_consecutives(cell_list: List, strip_col_letter: bool = True) -> str:
    """Takes a list of individual excel cells, returns a string summarising cell ranges for human readers.

    Args:
        cell_list (list of str or int): the excel cells, each expressed in the format "F1", or just "1" or 1.

        strip_col_letter (bool): whether or not the cell_list includes column letters that must be removed.

    Returns:
        out_string (str): the cell ranges, expressed using hyphens and commas.

    Examples:
        cell_list = [F1,F2,F4,F6,F7,F8], strip_col_letter=True
        out_string = "1-2, 4, 6-8"
    """
    # may be the case with comparative fails
    # that don't neatly relate to particular cells
    if len(cell_list) == 0:
        return ""

    if strip_col_letter:
        num_list = _strip_col_letter(cell_list)
    else:
        num_list = [int(val) for val in cell_list]

    num_list = sorted(num_list)
    assert len(num_list) == len(
        set(num_list)
    ), "I've found duplicate cell refs - did something go wrong with your groupby?"
    out_list = []
    for index, val in enumerate(num_list):

        # always keep first and last
        if index == 0 or index == (len(num_list) - 1):
            out_list.append(val)
        else:
            prev_val, next_val = num_list[index - 1], num_list[index + 1]
            # always keep isolates
            # and we know we'll never need to read them as ints for banding
            if (prev_val < val - 1) and (next_val > val + 1):
                out_list.append(str(val))
            # start of a band
            elif (prev_val != val - 1) and (next_val == val + 1):
                out_list.append(val)
            # end of a band
            elif (prev_val == val - 1) and (next_val != val + 1):
                prev_out = out_list.pop()
                out_list.append(str(prev_out) + "-" + str(val))

    # clean up, relying on the fact that isolates are already str
    if (
        (len(out_list) > 1)
        and isinstance(out_list[-2], int)
        and isinstance(out_list[-1], int)
    ):
        band_start = out_list.pop(-2)
        band_end = out_list.pop(-1)
        out_list.append(str(band_start) + "-" + str(band_end))

    out_list = [str(val) for val in out_list]
    out_string = (", ").join(out_list)

    return out_string


def _generate_template_order(template_map: Dict) -> Dict:
    """
    Assign a sort rank to each variable based on order of appearance in the template.

    Notes:
        This relies on the order of dictionary elements being consistent,
        which is guaranteed as of python 3.7 and semi-reliable as of 3.6.
    """
    template_order = {}
    order_marker = 0

    for key, inner_dict in template_map.items():
        id = (" | ").join([inner_dict["tab"], inner_dict["name"]])
        rank = order_marker
        template_order[id] = rank
        order_marker += 1

    return template_order


# TODO: *might* be able to make this cleverer by specifying a 3rd column to sort by next
# i.e. by multiplying out the order dict
def _sort_by_template_order(
    df: pd.DataFrame, template_order: Dict
) -> pd.DataFrame:
    """Sorts a df's rows based on tab and variable name."""

    if "concat_id" not in df.columns:
        df["concat_id"] = df["tab"] + " | " + df["name"]

    df = df.sort_values(by="concat_id", key=lambda x: x.map(template_order))
    df = df.drop("concat_id", axis=1)
    df = df.reset_index(drop=True)
    # actually sometimes misleading, I'll have a think about this
    # df.index.name = "template_order"

    return df


def _generate_check_order(vc: ModuleType = vc) -> List:
    """Inspects the code for basic validation checks to find the order for sorting lists of checks.

    Args:
        vc (python module) - automated_metadata_validation.validation._validation_checks

    Returns:
        just_names (list of str): a list of the names of each non-commented, non-internal
        function in the vc script, in the meaningful order in which they were written.
    """
    vc_text = inspect.getsource(vc)
    # all functions that don't start with # and aren't internal (_)
    func_matches = re.findall(pattern=r"(?<!# )def [^_].*\(", string=vc_text)
    # strip the def and (
    # which I could equally do with a more complicated regex probably
    just_names = [funcmatch[4:-1] for funcmatch in func_matches]

    return just_names


def _sort_by_check_order(df: pd.DataFrame, check_list: List) -> pd.DataFrame:
    """Sorts a df's *columns* based on validation check names."""

    not_check_cols = [
        colname for colname in df.columns if colname not in check_list
    ]
    check_cols = [colname for colname in check_list if colname in df.columns]

    col_order = not_check_cols + check_cols
    df = df[col_order]

    return df


def _group_and_format_cell_refs(
    df: pd.DataFrame, group_cols: List, template_order: Dict
) -> pd.DataFrame:
    """Groups a fails_df by the specified columns, finds the cell references belonging to each group, and expresses them as ranges.

    Args:
        df
        group_cols
        template_order

    Returns:
        group_df
    """
    group_df = df.groupby(group_cols, as_index=False).agg(
        {"cell_refs": lambda x: list(x)}
    )
    group_df["cell_refs"] = group_df["cell_refs"].apply(drop_consecutives)

    if ("tab" in group_df.columns) and ("name" in group_df.columns):
        group_df = _sort_by_template_order(group_df, template_order)

    return group_df


############################################
"""Functions for specific summary tables"""
############################################


def group_comparatives(
    full_df: pd.DataFrame, template_order: Dict
) -> pd.DataFrame:
    """Groups a fails_df so that each row details one instance of a failed comparative check.

    Args:
        full_df (pd.DataFrame): _description_
        template_order (Dict): _description_

    Returns:
        pd.DataFrame: _description_
    """

    comp_df = full_df.loc[full_df["fail_type"] == "comparative_fails"]
    comp_group_df = _group_and_format_cell_refs(
        comp_df, ["tab", "name", "reason", "col_ref"], template_order
    )

    return comp_group_df


def group_missing_values_by_variable(
    full_df: pd.DataFrame, template_order: Dict
) -> pd.DataFrame:
    """Groups a fails_df so that each row details the cells with missing values for a single variable.

    Args:
        full_df (_type_): _description_

    Returns:
        pd.DataFrame: _description_
    """
    miss_df = full_df.loc[full_df["reason"] == "missing_value. "]
    miss_group_df = _group_and_format_cell_refs(
        miss_df, ["tab", "name", "col_ref"], template_order
    )

    return miss_group_df


def group_by_check(
    full_df: pd.DataFrame, template_order: Dict
) -> pd.DataFrame:
    """Groups a fails_df so that each row details the cells of a single variable that have failed a particular hard or soft check.

    Args:
        full_df (pd.DataFrame): _description_
        template_order (Dict): _description_

    Returns:
        pd.DataFrame: _description_
    """

    non_comp_df = full_df.loc[
        full_df["fail_type"].isin(["hard_fails", "soft_fails"])
    ]
    check_group_df = _group_and_format_cell_refs(
        non_comp_df, ["tab", "name", "reason"], template_order
    )

    return check_group_df


def group_by_cell(full_df: pd.DataFrame, template_order: Dict) -> pd.DataFrame:
    """Groups a fails_df so that each row details the names of all hard and soft checks failed by a single cell.

    Args:
        full_df (pd.DataFrame): _description_
        template_order (Dict): _description_

    Returns:
        pd.DataFrame: _description_
    """

    non_comp_df = full_df.loc[
        full_df["fail_type"].isin(["hard_fails", "soft_fails"])
    ]
    cell_group_df = non_comp_df.groupby(
        ["tab", "name", "value", "cell_refs"], as_index=False
    ).agg({"reason": lambda x: "".join(x)})
    cell_group_df = _sort_by_template_order(cell_group_df, template_order)

    return cell_group_df


# In practice I don't think this one is going to be very useful,
# and I'm inclined to drop it from our outputs
# but we'll see how it fares with users.
def group_by_value(
    full_df: pd.DataFrame, template_order: Dict
) -> pd.DataFrame:
    """Groups a fails_df so that each row details a value appearing in a variable, all the cells that value appears in, and all the hard and soft checks that value fails.

    Args:
        full_df (pd.DataFrame): _description_
        template_order (Dict): _description_

    Returns:
        pd.DataFrame: _description_
    """

    non_comp_df = full_df.loc[
        full_df["fail_type"].isin(["hard_fails", "soft_fails"])
    ]
    value_group_df = non_comp_df.groupby(
        ["tab", "name", "value", "col_ref"], as_index=False
    ).agg(
        {"reason": lambda x: "".join(x), "cell_refs": lambda x: list(set(x))}
    )
    value_group_df["cell_refs"] = value_group_df["cell_refs"].apply(
        drop_consecutives
    )
    value_group_df = _sort_by_template_order(value_group_df, template_order)
    col_ref_col = value_group_df.pop("col_ref")
    value_group_df.insert(
        (len(value_group_df.columns) - 1), "col_ref", col_ref_col
    )

    return value_group_df


def pivot_and_percent_by_fail_reason(
    full_df: pd.DataFrame,
    row_count_df: pd.DataFrame,
    template_order: Dict,
    check_list: List,
) -> pd.DataFrame:
    """Produces a df where rows are tab & variable name combos and columns are fail %s for every check."""

    non_comp_df = full_df.loc[
        full_df["fail_type"].isin(["hard_fails", "soft_fails"])
    ]
    group_df = non_comp_df.groupby(
        ["tab", "name", "reason"], as_index=False
    ).agg({"cell_refs": "count"})
    pivot_df = group_df.pivot(
        index=["tab", "name"], columns="reason", values="cell_refs"
    ).reset_index()

    join_df = pd.merge(pivot_df, row_count_df, on="tab")
    join_df = join_df.loc[
        join_df["row_count"] != 0
    ]  # avoiding div/0 for Codes and Values
    join_df = join_df[
        ["tab", "name", "row_count"] + list(pivot_df.columns[2:])
    ]

    # TODO: will eventually want to denote checks not applied to particular variables
    # with N/As...
    # ...but for now, I'll mask with zeroes. Since we're doing fail %s, that'll still
    # be 0% fails!

    join_df = join_df.fillna(0)

    for colname in list(join_df.columns[3:]):
        join_df[colname] = np.round(
            (join_df[colname] / join_df["row_count"]) * 100, 1
        )

    out_df = _sort_by_check_order(join_df, check_list)
    out_df = _sort_by_template_order(out_df, template_order)

    return out_df


def make_short_summary_by_variable(
    full_df: pd.DataFrame, row_count_df: pd.DataFrame, template_order
) -> pd.DataFrame:
    """Produces a df where rows are tab & variable name combos and the columns list the % of records that are missing or failed at least one check.

    Args:
        full_df (pd.DataFrame): _description_
        row_count_df (pd.DataFrame): _description_
        template_order (_type_): _description_

    Returns:
        pd.DataFrame: _description_
    """

    # TODO: might want to also distinguish between hard and soft fails
    # TODO: might be able to link in "does at least one comparative check fail refer to this column"?

    cell_df = group_by_cell(
        full_df, template_order
    )  # currently sets aside comparative fails

    # if it's missing, it will never have any other fail reasons, because they're moot
    miss_df = cell_df.loc[cell_df["reason"] == "missing_value. "]
    miss_count_df = miss_df.groupby(["tab", "name"], as_index=False).agg(
        {"cell_refs": "count"}
    )
    miss_count_df = miss_count_df.rename(
        {"cell_refs": "missing values"}, axis=1
    )

    fail_df = cell_df.loc[cell_df["reason"] != "missing_value. "]
    fail_count_df = fail_df.groupby(["tab", "name"], as_index=False).agg(
        {"cell_refs": "count"}
    )
    fail_count_df = fail_count_df.rename(
        {"cell_refs": "failed at least one check"}, axis=1
    )

    # TODO: may also want to add rows for variables that didn't fail at all,
    # so that it's clear that they were checked and didn't fail!
    var_df = (
        full_df[["tab", "name"]]
        .drop_duplicates()
        .sort_values(["tab", "name"])
        .reset_index(drop=True)
    )

    join1_df = pd.merge(var_df, row_count_df, on="tab", how="left")
    join1_df = join1_df.fillna(1)  # vars on cellwise tabs
    join1_df = join1_df.rename({"row_count": "total records"}, axis=1)

    join2_df = pd.merge(
        join1_df, miss_count_df, on=["tab", "name"], how="left"
    )
    join3_df = pd.merge(
        join2_df, fail_count_df, on=["tab", "name"], how="left"
    )
    join3_df = join3_df.fillna(
        0
    )  # i.e. where there's missing values but not failed checks, or vice versa

    join3_df["no issues detected"] = (
        join3_df["total records"]
        - join3_df["missing values"]
        - join3_df["failed at least one check"]
    )

    percent_df = join3_df
    for colname in [
        "missing values",
        "failed at least one check",
        "no issues detected",
    ]:
        percent_df[colname] = np.round(
            (percent_df[colname] / percent_df["total records"]) * 100, 1
        )
        percent_df = percent_df.rename({colname: colname + " %"})

    out_df = _sort_by_template_order(percent_df, template_order)

    return out_df


##################################################
"""Functions related to writing the excel files"""
##################################################


def _fit_and_wrap(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """Approximates a sensible width and wrap threshold based on number of characters in cells."""

    for colname in df.columns:

        max_cell_len = max(df[colname].astype(str).str.len())
        max_len = min(200, max(max_cell_len, len(colname)))

        # whether you offset by 1 or 2 depends on whether you're outputting the index
        number_index = list(df.columns).index(colname) + 2
        letter_index = get_column_letter(number_index)

        wb[sheet_name].column_dimensions[letter_index].width = max_len

        if max_len == 200:
            for cell in wb[sheet_name][letter_index]:
                cell.alignment = Alignment(wrap_text=True)


def _split_by_mandatory(
    fails_df: pd.DataFrame, template_map: Dict
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Uses the template map to split processor.fails_df in twain"""
    mand_data_list = []
    for key, inner_dict in template_map.items():
        mand_data_list.append(
            [inner_dict["tab"], inner_dict["name"], inner_dict["mandatory"]]
        )

    mand_df = pd.DataFrame(
        data=mand_data_list, columns=["tab", "name", "mandatory"]
    )

    join_df = pd.merge(fails_df, mand_df, on=["tab", "name"], how="left")

    mand_df = join_df.loc[join_df["mandatory"]]
    non_mand_df = join_df.loc[~join_df["mandatory"]]

    assert len(mand_df) + len(non_mand_df) == len(
        fails_df
    ), "Looks like something's gone wrong when joining the mandatory tags!"

    return mand_df, non_mand_df


def _make_save_path(
    processor, save_folder: str, mode: str = "main_report"
) -> str:
    """Produces names and filepaths for output files, based on the name of the input file.

    Args:
        processor (MetadataProcessor): _description_
        save_folder (str): _description_
        mode (str, optional): _description_. Defaults to "main_report".

    Raises:
        ValueError: _description_

    Returns:
        str: _description_
    """

    save_name = os.path.basename(processor.md_filepath)
    save_name = os.path.splitext(save_name)[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    if mode == "main_report":
        save_name = f"{save_name}_VALIDATION_REPORT_{timestamp}.xlsx"
    elif mode == "comment_copy":
        save_name = f"COMMENT_COPY_DO_NOT_SUBMIT_{save_name}_{timestamp}.xlsx"
    else:
        raise ValueError(
            "Please specify mode = 'main_report' or 'comment_copy'"
        )

    save_path = os.path.join(save_folder, save_name)

    return save_path


def _create_output_template() -> Workbook:
    """Creates a workbook with a guidance sheet containing user documentation.

    Notes:
        This is necessary for packaging, since Pypi doesn't like including .xlsx files.
    """

    wb = openpyxl.Workbook()
    wb["Sheet"].title = "Guidance"

    cell_text = {
        "B2": "Automated Metadata Validation Tool",
        "B3": "v0.1.0",
        "B4": "Welcome to the Automated Metadata Validation output report!",
        "B5": "This sheet is a placeholder for now, but will eventually contain further information about how to use this report.",
        "B7": "If you have an ONS Digital github account, you can also view the main readme here:",
        "B8": "https://github.com/ONSdigital/automated-metadata-validation",
        "B10": "Metadata variables are either mandatory or optional.",
        "B11": 'Validation checks are considered to be "hard", "soft", or "comparative".',
        "B13": "Hard checks are conditions that can be conclusively measured automatically. Failing a hard check means that something is definitely wrong and needs changing. This also means that hard check fails will usually also cause an ingest failure if untreated, since the ingest process also has fixed expectations about machine-readable content and formats.",
        "B15": "Soft checks are checks that require inspection, but not necessarily action, if they fail. Either they relate to style recommendations  that aren't strict requirements, or they involve checking something that can't be perfectly measured automatically. For example, we may expect a certain style of response most of the time, but there may be corner cases where unusual answers are still acceptable and correct.",
        "B17": "Comparative checks involve more than one cell value at a time. For example, a column of table names might require that each name be unique within that column. Or, for consistency, a table name appearing on one sheet might be required to also appear on a list of tables from a previous sheet.",
        "B19": "Sheet",
        "B20": "Short % overview",
        "B21": "Long % overview",
        "B22": "Missing values",
        "B23": "Fails by cell",
        "B24": "Fails by check",
        "B25": "Fails by value",
        "B26": "Comparative checks",
        "B27": "Non-mandatory fails by cell",
        "C19": "Variables",
        "C20": "All mandatory",
        "C21": "All mandatory",
        "C22": "All mandatory",
        "C23": "All mandatory",
        "C24": "All mandatory",
        "C25": "All mandatory",
        "C26": "Comparative only",
        "C27": "Non-mandatory only",
        "D19": "Description",
        "D20": "each row is a tab & variable name combination; columns list the % of records that are missing or failed at least one check.",
        "D21": "each row is a tab & variable name combination; columns are fail %s for every check.",
        "D22": "each row details the cells with missing values for a single variable.",
        "D23": "each row details the names of all hard and soft checks failed by a single cell.",
        "D24": "each row details the cells of a single variable that have failed a particular hard or soft check.",
        "D25": "each row details a value appearing in a variable, all the cells that value appears in, and all the hard and soft checks that value fails. NOTE: this view is experimental and has some known bugs with cell ranges.",
        "D26": "each row details one instance of a failed comparative check.",
        "D27": "each row details the names of all hard and soft checks failed by a single cell, including missing values. Non-mandatory variables only.",
    }

    for cell, text in cell_text.items():
        wb["Guidance"][cell].value = text

    """Formatting"""

    # These are simply known from having previously mocked up the sheet in excel
    wb["Guidance"].column_dimensions["B"].width = 30
    wb["Guidance"].column_dimensions["C"].width = 20
    wb["Guidance"].column_dimensions["D"].width = 120

    for row in ["13", "15", "17"]:
        wb["Guidance"].row_dimensions[row].height = 45
        wb["Guidance"].merge_cells(
            start_row=row, start_column=2, end_row=row, end_column=4
        )

    for cell in ["B13", "B15", "B17", "D25"]:
        wb["Guidance"][cell].alignment = Alignment(wrap_text=True)

    # https://stackoverflow.com/questions/13650059/apply-borders-to-all-cells-in-a-range-with-openpyxl
    thin = Side(border_style="thin", color="000000")
    for row in wb["Guidance"]["B19:D27"]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    return wb


def write_output_report(processor, save_folder: str) -> openpyxl.Workbook:
    """Takes a MetadataProcessor object containing a df of failed checks and other metametadata, produces a validation report in excel."""

    fails_df = processor.fails_df
    fails_df["col_ref"] = fails_df["cell_refs"].astype(str).str[0]
    # temporary - some issues with duplication earlier in processing
    fails_df = fails_df.drop_duplicates()

    mand_df, non_mand_df = _split_by_mandatory(
        fails_df, processor.template_map
    )
    # may ask for row counts as df in next iteration
    row_count_df = pd.DataFrame(
        data=list(processor.tab_lens.items()), columns=["tab", "row_count"]
    )
    template_order = _generate_template_order(processor.template_map)
    check_order = _generate_check_order(vc)

    wb = _create_output_template()

    out_dfs = {}

    out_dfs["Short % overview"] = make_short_summary_by_variable(
        mand_df, row_count_df, template_order
    )
    out_dfs["Long % overview"] = pivot_and_percent_by_fail_reason(
        mand_df, row_count_df, template_order, check_order
    )
    out_dfs["Missing values"] = group_missing_values_by_variable(
        mand_df, template_order
    )
    out_dfs["Fails by cell"] = group_by_cell(mand_df, template_order)
    out_dfs["Fails by check"] = group_by_check(mand_df, template_order)
    out_dfs["Fails by value"] = group_by_value(mand_df, template_order)
    out_dfs["Comparative checks"] = group_comparatives(mand_df, template_order)
    out_dfs["Non mandatory fails by cell"] = group_by_cell(
        non_mand_df, template_order
    )

    # TODO: fancier sheet formatting
    # - as table with borders
    # - heatmaps / colour coding, e.g. for the % overviews

    for sheet_name, df in out_dfs.items():
        wb.create_sheet(sheet_name)
        # We now do want to keep the index, because it's often a meaningful sort order
        for row in dataframe_to_rows(df, index=True, header=True):
            wb[sheet_name].append(row)

        if not df.empty:
            _fit_and_wrap(wb, sheet_name, df)

    if processor.save_report:
        save_path = _make_save_path(processor, save_folder, mode="main_report")
        wb.save(save_path)
        print(f"Saved to {save_path}!")

    return wb


def create_copy_with_comments(
    processor, save_folder: str
) -> openpyxl.Workbook:
    """Uses the df of check fails to add comments and highlighting to a copy of the original template."""

    fails_df = processor.fails_df.drop_duplicates()
    template_order = _generate_template_order(processor.template_map)
    cellwise_df = group_by_cell(fails_df, template_order)

    metadata_wb = openpyxl.load_workbook(processor.md_filepath)

    template_order = _generate_template_order(processor.template_map)

    # I don't think excel 'notes' actually display this
    author = "Automated Metadata Validation Tool"

    yellow_fill = PatternFill(
        start_color="00FFFF00", end_color="00FFFF00", fill_type="solid"
    )

    for index, df_row in cellwise_df.iterrows():
        sheet = metadata_wb[df_row["tab"]]
        cell = sheet[df_row["cell_refs"]]
        comment = cell.comment
        com_text = df_row["reason"].replace("_", " ")  # slightly more readable
        cell.comment = Comment(com_text, author)
        # print(sheet, cell, df_row["value"], comment) #testing
        cell.fill = yellow_fill

    # Unfortunately, attempts at protection sheets don't work because of the existing protection
    metadata_wb["Change history"][
        "D1"
    ].value = "Validation comments doc - NOT FOR INGESTION"
    metadata_wb["Dataset Resource"][
        "A1"
    ].value = "Validation comments doc - NOT FOR INGESTION"
    metadata_wb.create_sheet("Important notice")
    metadata_wb.active = metadata_wb["Important notice"]
    metadata_wb["Important notice"][
        "A1"
    ].value = "DO NOT edit and submit this workbook."
    metadata_wb["Important notice"][
        "A2"
    ].value = "It is a commented copy for reference only."
    metadata_wb["Important notice"][
        "A3"
    ].value = "If you submit this file, it will not be accepted."
    metadata_wb["Important notice"][
        "A4"
    ].value = (
        "Please reopen and edit the original, using this file as a guide."
    )

    save_path = _make_save_path(processor, save_folder, mode="comment_copy")
    metadata_wb.save(save_path)
    print(f"Saved to {save_path}!")

    return metadata_wb
