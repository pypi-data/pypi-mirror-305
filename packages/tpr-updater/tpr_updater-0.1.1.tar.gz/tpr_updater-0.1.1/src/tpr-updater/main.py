import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import filedialog, messagebox
import win32com.client as win32


def main():
    
    print("Please select the TPR file to process")

    # Load the Excel file using openpyxl (preserves formatting)
    file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])

    # if the file is an xls file, convert it to xlsx
    if file_path.endswith('.xls'):
        file_path = convert_xls_to_xlsx(file_path)
    elif not file_path:
        messagebox.showerror("Error", "No file selected. Exiting program.")
        exit()
    else:
        messagebox.showerror("Error", "File is not an xls or xlsx file. Exiting program.")
        exit()

    process_tpr_file(file_path)


def convert_xls_to_xlsx(xls_file):
    if not os.path.exists(xls_file):
        print(f"File not found: {xls_file}")
        return

    
    file_name = xls_file.split('.')[0]
    file_dir = os.path.dirname(xls_file)

    xlsx_file = os.path.join(file_dir, file_name + '.xlsx')
    print(f"Converting {file_name} to {xlsx_file}")
    
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(xls_file)

        if xlsx_file:
            # Normalize the path to use Windows-style separators
            xlsx_file = os.path.normpath(xlsx_file)

            # Close any open workbook with the same name
            for open_wb in excel.Workbooks:
                if open_wb.FullName.lower() == xlsx_file.lower():
                    open_wb.Close(SaveChanges=False)

            try:
                wb.SaveAs(xlsx_file, FileFormat=51)
                print(f"Conversion complete. File saved as {xlsx_file}")
            except Exception as save_error:
                print(f"Error saving file: {str(save_error)}")
                # Try an alternative save method
                try:
                    temp_path = os.path.join(os.path.dirname(xlsx_file), "temp_" + os.path.basename(xlsx_file))
                    wb.SaveAs(temp_path, FileFormat=51)
                    os.replace(temp_path, xlsx_file)
                    print(f"Conversion complete using alternative method. File saved as {xlsx_file}")
                except Exception as alt_save_error:
                    print(f"Alternative save method also failed: {str(alt_save_error)}")
        else:
            print("Save operation cancelled.")

        wb.Close()

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        import traceback
        print(traceback.format_exc())

    finally:
        excel.Quit()

    return xlsx_file

def process_tpr_file(file_path):

    print("Processing file: ", file_path)

    book = load_workbook(file_path)
    sheet_name = 'TPR Items'  
    sheet = book[sheet_name]

    # Load the data into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=[5, 6], engine='openpyxl')

    # Combine the multi-row headers into a single row
    df.columns = [' '.join(col).strip() for col in df.columns.values]

    # Perform your data modifications in pandas
    for index, row in df.iterrows():
        # Update TPR SRP if SRP is not 0 and less than TPR SRP
        if row['SRP SRP'] != 0 and row['SRP SRP'] < row['TPR SRP']:
            df.loc[index, 'SRP SRP'] = row['TPR SRP']
            # Mark this cell for formatting later
            sheet.cell(row=index+8, column=df.columns.get_loc('SRP SRP')+1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            print(f"Updated row {index+8}, column 'SRP SRP' from {row['SRP SRP']} to {row['TPR SRP']}")

    # Write the modified DataFrame back to the same sheet
    for r_idx, row in enumerate(df.values, 8):  # Start from row 8 (7 + 1 for 0-indexing)
        for c_idx, value in enumerate(row, 1):  # Start from column 1
            sheet.cell(row=r_idx, column=c_idx, value=value)

    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=os.path.basename(file_path))
    if output_file_path:
        book.save(output_file_path)
        print(f"Updated data saved to {output_file_path}")
        # messagebox.showinfo("Success", f"Updated data saved to {output_file_path}")


if __name__ == "__main__":
    main()