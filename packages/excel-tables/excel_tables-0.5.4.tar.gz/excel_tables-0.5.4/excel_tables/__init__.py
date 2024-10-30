#!/usr/bin/env python3
"""
Excel Functions
(C) Laurent Franceschetti (2024)
"""

import os
from collections.abc import Callable
from copy import copy
from dateutil.parser import parse as date_parse
from collections import namedtuple
from typing import List, get_type_hints
from functools import partial
import sqlite3
import inspect


from pydantic.dataclasses import dataclass
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, colors
from rich.console import Console, Capture
from rich.table import Table

from babel.numbers import format_decimal
from babel.dates import format_date


from .excel_util import (get_col_widths, expand_name, get_wks, 
                         get_column_id, to_argb, get_font_color,
                         xl_strftime, open_file)
# replaces 'from xlsxwriter.utility import xl_col_to_name':
from .excel_util import xl_col_to_name
from .df_util import df_columns, apply_to_column, convert_ISO_dates




# --------------------------------
# PARAMETERS
# --------------------------------

DEFAULT_FONT_NAME = 'Arial'



# colors of Excel
HEADER_COLOR = "#002966"
TEXT_COLOR_ERROR = "red"
TEXT_COLOR_WARNING = "#CD853F"  # peru
FILL_ALTERNATE = '#EEEEEE' # Light grey
FILL_EMPHASIS = '#FFFEC8'


# Numeric formats (default: English)
# Do not change, as they also apply to Babel
FORMAT_INT   = "#,##0"
FORMAT_FLOAT = "#,##0.00"
FORMAT_PERC = "0.0%"
FORMAT_DATE = "YYYY-MM-DD" # ISO, best for sorting, etc.



def convert_date(s):
    "Convert string date to European date"
    if isinstance(s, str):
        return date_parse(s, dayfirst=True)
    else:
        return s
NUM_MONTH_DATE = (convert_date, "mm-yyyy")
NUM_DATE =   (convert_date, "dd-mm-yyyy")




# for OpenPyxl (reports, etc.)
# FORMAT_TITLE_FONT = Font(bold=True, color=colors.WHITE)
FORMAT_TITLE_FILL = PatternFill(start_color=to_argb(HEADER_COLOR),
                                    fill_type='solid')
FORMAT_FILL_ALTERNATE = PatternFill(start_color=to_argb(FILL_ALTERNATE),
                                    fill_type='solid')
FORMAT_FILL_EMPHASIS = PatternFill(start_color=to_argb(FILL_EMPHASIS),
                                    fill_type='solid')
FORMAT_TITLE_ALIGNMENT = Alignment(vertical='top', 
                                   horizontal='center', wrap_text=True)

SheetImage = namedtuple('SheetImage', 'filename position')


# --------------------------------
# Reading object for an Excel file (import)
# --------------------------------
@dataclass(config=dict(arbitrary_types_allowed=True))
class ExcelDB(object):
    """
    Class for querying one or more Excel files with SQL
    """

    "The Excel file from which to read (optional)"
    filename: str | None = None
    
    "The (main) database file (default: in memory)"
    db_filename: str = ":memory:"

    "The data connection (opens at creation time)"
    conn: sqlite3.Connection = None


    def __post_init__(self):
        self.conn = sqlite3.connect(self.db_filename)
        if self.filename:
            self.import_file(self.filename)

    @staticmethod
    def extract_all_tabs(filename:str) -> list[tuple]:
        """
        Read an Excel file and return a list of tuples
        (tab, dataframe)
        """
        xlsx = pd.ExcelFile(filename)
        # Extract all sheet names
        sheet_names = xlsx.sheet_names
        # Create a list of dataframes for each sheet
        tabs = [(sheet, xlsx.parse(sheet)) 
                for sheet in sheet_names]
        
        return tabs


    def load(self, tablename: str, df:pd.DataFrame, replace:bool=True):
        """
        Load a Pandas dataframe (from any source) into the database,
        with the tablename.
        By default, it replaces the table.
        """
        if replace:
            if_exists='replace'
        else:
            if_exists='fail'
        df.to_sql(tablename, self.conn, if_exists=if_exists, index=False)

    def import_file(self, filename:str, replace:bool=True):
        """
        Import all tabs of an Excel spreadsheet into the SQLite database

        Every worksheet table (tab) becomes a table in the SQL database.

        An index with the same name replaces the previous one.
        """
        tabs = self.extract_all_tabs(filename)

        for tablename, df in tabs:
            # print(f"{tablename}")
            self.load(tablename, df, replace=replace)

    def load_wks(self, filename:str, sheetname:str|int, tablename:str=None,
                   replace:bool=True):
        "Import a worksheet from an Excel file; you can use the number"
        xl = pd.ExcelFile(filename)
        df = xl.parse(sheetname)
        # if it's an int we need to get the actual name:
        if isinstance(sheetname, int):
            sheetname = xl[sheetname]
        tablename = tablename or sheetname
        self.load(tablename, df, replace=replace)

    # -------------------------------------------
    # Query
    # -------------------------------------------
    def query(self, query:str, params:list|dict=None) -> pd.DataFrame:
        """
        Make an SQL query on against the database

        params can be a list or dictionary according to the sqlite
        syntax used (by default uses the caller's context as dictionary).
        See: https://docs.python.org/3/library/sqlite3.html#how-to-use-placeholders-to-bind-values-in-sql-queries
        
        Dates stored as ISO strings (as typically done) by Pandas
        are corrected.
        """
        if not params:
            caller_frame = inspect.currentframe().f_back
            params = caller_frame.f_locals

        df = pd.read_sql_query(query, self.conn, params=params)
        convert_ISO_dates(df)
        return df

    
    def get_values(self, query:str, params:list|dict=None) -> list:
        """
        Get a list of values from the database (one column query)

        NOTE
        ----
        This returns a list; good for most business purposes,
        but not for large queries.
        """
        if not params:
            caller_frame = inspect.currentframe().f_back
            params = caller_frame.f_locals
        df = self.query(query, params)
        return df.iloc[:, 0].tolist()
    
    def get_value(self, query:str,
                  params:list|dict=None) -> str | int | float:
        "Get a value from a database, typically for a `count(*)`"
        if not params:
            caller_frame = inspect.currentframe().f_back
            params = caller_frame.f_locals
        return self.get_values(query, params)[0]
    
    @property
    def tables(self) -> list[str]:
        "The list of tables loaded"
        query = "SELECT name FROM sqlite_master WHERE type='table';"
        return self.get_values(query)
    

    def table(self, id:str|int=0) -> pd.DataFrame:
        "Gets the table by its number or name (default: first available)"
        if isinstance(id, int):
            tablename = self.tables[id]
        else:
            # it's the name
            tablename = id
        return self.query(f"SELECT * FROM [{tablename}]")
    
    
    # -------------------------------------------
    # Close or delete data
    # -------------------------------------------


    def drop(self, tablename:str):
        "Drop a table from the database"
        if not tablename in self.tables:
            raise NameError(f"Cannot delete '{tablename}' (non-existent)")
        cursor = self.conn.cursor()
        cursor.execute(f"DROP TABLE {tablename}")

    def close(self):
        "Close the database"
        self.conn.close()

    def __del__(self):
        self.close()





# --------------------------------
# Low-level export builders
# --------------------------------

def format_xl_report(filename: str, 
                   tabs:dict=None,
                   font_name:str=DEFAULT_FONT_NAME,
                   header_colors:dict|str=HEADER_COLOR,
                   number_formats: dict = {},
                   emphasizes: dict = {},
                   tab_colors: dict = {},
                   images: dict = {},
                   debug:bool=False) -> openpyxl.Workbook:
    """
    Format an Excel report (with pre-existing tables)

    Uses openpyxl

    Arguments
    ---------
    filename: filename
    tabs: subset of the tabs to be formatted (index or string)
      Regex must be preced by '//'
      By default, ALL tabs will get default formatting.
    number_formats: a dictionary of spreadsheet tabs (index or string),
      containing a list of duplets: column (name or index) and
      format (name or index). 
      For number_style indexes, see https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
    header_colors: the color to be used for the first row;
        can be either a string (valid for all)
        or a dictionary of spreadsheet tabs
        (name or index)
    emphasizes: A dictionary of spreadsheet tabs (index or string),
        containing a dictionary of predicate functions 
        for a row (= list of cell values), 
        to determine whether it should be emphasized.
    tab_colors: a dictionary of spreadsheet tabs,
        with color of the tab (if one wants to give it)
    images: a dictionary of spreadsheet tabs,
        with an image

    Additional info
    ---------------  
    The tab and column indexes are counted from zero (Python standard).
    
    For number_style indexes, see https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html 

    Example of numeric format:
        {'first_tab: [('foo': '#,##0'), (5, 2)], 
          2:          [(2: ''mm-dd-yy')]}

    Returns
    -------
    The openpyxl workbook.
    """
    if not os.path.exists(filename):
        raise FileNotFoundError(f"Spreadsheet '{filename}' not found.")
    
    
    wb = openpyxl.load_workbook(filename)
    wb.iso_dates = True

    format_title_fill = FORMAT_TITLE_FILL
    if debug:
        print("FONT:", font_name)
        print("Header colors:", header_colors)
    font = Font(name=font_name)
    # by default, all of them
    if tabs is None:
        tabs = wb.sheetnames

    for tab in tabs:
        # determine header colors
        if header_colors:
            if isinstance(header_colors, str):
                header_color = header_colors
            else:
                header_color=header_colors.get(tab, HEADER_COLOR)
            format_title_fill = PatternFill(start_color=to_argb(header_color),
                                        fill_type='solid')
            if get_font_color(header_color) == 'white':
                header_font_color = colors.WHITE
            else:
                header_font_color = colors.BLACK
        emphasize = emphasizes.get(tab) or (lambda x: False)
        
        # get the worksheet:
        ws = get_wks(wb, tab)
        header = []
        rows = []
        # set the tab color
        tab_color = tab_colors.get(tab)
        if tab_color:
            # accept a dash in front of it
            ws.sheet_properties.tabColor = to_argb(tab_color)
        # define the header (first row) format:
        for i in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=i)
            header.append(cell.value)
            #cell.font = font
            cell.font = Font(name=font_name, bold=True, color=header_font_color)
            # cell.font = FORMAT_TITLE_FONT
            cell.fill = format_title_fill
            cell.alignment = FORMAT_TITLE_ALIGNMENT
        # define the rest:
        for i in range(2, ws.max_row + 1):
            # NOTE: Important to go to the highest row no + 1
            row = []
            even_row = (i % 2 == 0)
            for j in range(1, ws.max_column + 1):
                cell = ws.cell(row=i, column=j)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                row.append(cell.value)
                cell.font = font
                if even_row:
                    # alternate row
                    cell.fill = FORMAT_FILL_ALTERNATE 
            # emphasize predicate
            if emphasize(row):
                for j in range(1, ws.max_column + 1):
                    cell = ws.cell(row=i, column=j)
                    # cell.font = Font(name=font_name, bold=True)
                    cell.fill = FORMAT_FILL_EMPHASIS
            rows.append(row)
        # insert an image
        image = images.get(tab)
        if image:
            if not os.path.exists(image.filename):
                err_msg = f"Image '{os.path.basename(image.filename)}' does not exist"
                raise FileExistsError(err_msg)
            img_obj = openpyxl.drawing.image.Image(image.filename)
            img_obj.anchor = image.position
            ws.add_image(img_obj)

        # ------------------------
        # set the numeric formats 
        # ------------------------
        number_formats_col = number_formats.get(tab)

        if number_formats_col:
            if debug:
                print(f"- Numeric formats for {tab}", number_formats_col)
            for col, number_format in number_formats_col.items():
                # expand into a list (in case of regex)
                assert isinstance(number_format, str),  f"Tab '{tab}', col '{col}': Format not a str: {number_format}"
                exp_cols = expand_name(col, header, desc='column')
                if debug:
                    print("  Expanded:", exp_cols)
                for exp_col in exp_cols:
                    column_id = get_column_id(header, exp_col)
                    column_no = get_column_id(header, exp_col, return_int=True)
                    if debug:
                        print(f"  - Num format {exp_col}({column_no}) ({column_id}) => {number_format}")
                    ws.column_dimensions[column_id].number_format = number_format
                    for row_no in range(2, ws.max_row + 1):
                        # print(f"Apply to {row_no}/{column_no}")
                        ws.cell(column=column_no, 
                                row=row_no).number_format = number_format

        # define the column width
        column_widths = get_col_widths(header, rows) 
        for i, width in enumerate(column_widths):
            # get the ASCII value
            col = xl_col_to_name(i)
            ws.column_dimensions[col].width = width

        # freeze first row
        ws.freeze_panes = ws['A2']
        # autofilter:
        ws.auto_filter.ref = ws.dimensions
        # remove grid lines
        ws.sheet_view.showGridLines = False
    wb.save(filename)
    return wb






def format_text_table(df: pd.DataFrame, header_color:str,
                       printing:bool=True):
    """
    Prepare a rich text table for the console,
    with sensible (though highly opinionated) defaults.

    The idea is to emulate more or less the aspect on Excel,
    minus specific numeric formats.  
    """
    # we assume this for console printing of numbers
    LOCALE = 'en_us'
    # prepare a new dataframe for printing:
    col_specs = df_columns(df)
    assert df is not None
    output_df = df.copy(deep=False)
    for col in output_df:
        spec = col_specs[col]
        # print(col, spec)
        if spec == 'int':
            apply_to_column(output_df, col, 
                partial(format_decimal, format=FORMAT_INT, 
                        locale=LOCALE))
        elif spec == 'float':
            apply_to_column(output_df, col, 
                partial(format_decimal, format=FORMAT_FLOAT, 
                        locale=LOCALE))
        elif spec == 'perc':
            apply_to_column(output_df, col, 
                partial(format_decimal, format=FORMAT_PERC, 
                        locale=LOCALE))
        elif spec == 'date':
            # We use the ISO norm which is 'transatlantic'
            apply_to_column(output_df, col, 
                partial(xl_strftime, style=FORMAT_DATE))


    DEFAULT_STYLE = "black on white"
    console = Console()
    font_color = get_font_color(header_color)
    table = Table(show_header=True, 
                    header_style=f"bold {font_color} on {header_color}",
                    style=DEFAULT_STYLE,
                    padding=(0, 1))

    # Add columns
    for col in output_df.columns:
        justify = 'left'
        if col_specs[col] != 'str':
            justify = 'right'
        table.add_column(col, style=DEFAULT_STYLE, justify=justify)

    # Add rows
    for _, row in output_df.iterrows():
        table.add_row(*[str(item) for item in row])
    if printing:
        console.print(table)
        return None
    else:
        with Capture(console) as capture:
            console.print(table)
        return capture.get()

# --------------------------------
# Create Excel workbook (from pandas)
# --------------------------------

@dataclass(config=dict(arbitrary_types_allowed=True))
class Worksheet(object):
    """
    Dataframe as an Excel Worksheet
    
    Arguments
    ---------
    sheet_name: the name of the sheet (tab) to be created
    df: dataframe to be stored
    num_formats: dictionary (column, col_format) 
        or list of duplets (column, col_format)
        A format can be either: 
            - an Excel numeric format
            - a duplet (callable, numeric_format)
    emphasize: A predicate function for a row 
        (= list of values), 
        to determine whether it should be emphasized.
    tab_color: color of the tab for this sheet
    image: filename for a file to insert
    image_position: position in the sheet

    Example:
    --------
    Format column foo; apply transformation to column bar, and format
    num_format = {'foo': "#.0", 
                    'bar': (lambda s: int(s) + 1, "#.0")}

    Note
    ----
    All numeric column names (e.g. for years) are automatically 
    converted to string.
    """

    "The name of the sheet"
    sheet_name:str

    "The dataframe representing the data"
    df: pd.DataFrame

    "The numeric formats (special instructions, on top of defaults)"
    num_formats:dict| list | None = None

    "The function for emphasizing the rows"
    emphasize: Callable | None = None

    "The color of the tab in the Excel window"
    tab_color: str = None

    "Filename for an image to be displayed"
    image_filename: str = ''

    "On which cell the image will be positioned"
    image_position: str = 'B2'

    "Header (first line) color"
    header_color: str = HEADER_COLOR



    def __post_init__(self):
        "Actions after the init"
        if not self.num_formats:
            self.num_formats = []
        self.rebuild()

    def rebuild(self):
        "Rebuilds the worksheet"
        sheet_name = self.sheet_name
        if not sheet_name:
            raise ValueError(f"Sheet name cannot be empty")

        if self.df is not None:
            df_s = self.df.copy()
            # make sure all columns of dataframe are string:
            df_s.columns = df_s.columns.map(str) # all string cols
            self.df = df_s

            
            # make sure spec is as duplet
            if isinstance(self.num_formats, dict):
                self.num_formats = [(key, value) for key, value 
                                        in self.num_formats.items()]
                # print("NEW NUM_FORMAT", num_formats)
            # apply modifications requested on column:
            nf = []
            header = df_s.head()

            # apply numerical formats
            for col, col_format in self.num_formats:
                if isinstance(col_format, str):
                    # normal (don't expand)
                    nf.append((col, col_format))
                else:
                    # first item is a callable, apply it
                    call, num_format = col_format
                    # expand columns now
                    exp_cols = expand_name(col, header, desc='column')
                    for c in exp_cols:
                        df_s[c] = df_s[c].map(call)
                        assert isinstance(num_format, str)
                        nf.append((c, num_format))
            self.num_formats = nf

    @property
    def columns(self) -> dict:
        "Return the type of columns"
        return df_columns(self.df)




    def __str__(self):
        "Printable output"
        return self.rich_print(printing=False)



@dataclass(config=dict(arbitrary_types_allowed=True))
class ExcelReport(object):
    """
    Object used to create an Excel report from Pandas dataframes.

    1. Create the report (filename and general parameters)
    2. Append the worksheet
    3. Save the report (will create the file)

    This does not modify an existing file.
    """

    "Filename fo the file"
    filename: str


    "Worksheets that will be part of the file (1 dataframe per worksheet)"
    worksheets: List[Worksheet] | None = None

    "The reference font"
    font_name: str = DEFAULT_FONT_NAME

    "Color of the header"
    header_color:str=HEADER_COLOR

    """
    Dataframe for first worksheet, if there is only one.
    In that case, the file is immediately saved.
    """
    df: pd.DataFrame | None = None

    "Name of first worksheet (tab, if there is only one)"
    tabname:str = 'Main'

    "The numeric formats (if there is one worksheet)"
    num_formats:dict | list | None = None

    "The function for emphasizing the rows  (if there is one worksheet)"
    emphasize: Callable | None = None


    """
    Automatically save a file once a worksheet is appended.
    """
    auto_save:bool=False

    "Numeric format"
    format_int: str = FORMAT_INT

    "Numeric format"
    format_float: str = FORMAT_FLOAT

    "Percent format"
    format_perc:str = FORMAT_PERC

    "Date format (no hours and minutes)"
    format_date:str = FORMAT_DATE

    "Show debug info"
    debug:bool=False

    "The openpyxl Excel object (report)"
    workbook: openpyxl.Workbook = None


    def __post_init__(self):
        if not self.filename:
            raise ValueError("Filename for Excel report may not be empty")
        self.worksheets = []
        if self.df is not None:
            wks = Worksheet(self.tabname, self.df,
                            emphasize=self.emphasize)
            self.append(wks)
            self.save()


    def append(self, worksheet:Worksheet):
        "Append a worksheet with a specific name"
        self.worksheets.append(worksheet)
        if self.auto_save:
            self.save()

    def add_sheet(self, sheet_name:str, df: pd.DataFrame, 
                  num_formats:dict|list=None,
                  emphasize: Callable=None,
                  *args, **kwargs):
        """
        Add a worksheet (same arguments as for the Worksheet object).

        Typically:

        self.add_sheet('foo', df)
        """
        worksheet = Worksheet(sheet_name, df, num_formats=num_formats,
                              emphasize=emphasize,
                              *args, **kwargs)
        self.append(worksheet)
        return worksheet   


    # -------------------------------------------
    # These are utility functions to build
    #  the final call to format_xl_report
    # -------------------------------------------
    @property
    def tabs(self) -> list[str]:
        "The list of worksheet tables"
        return [wks.sheet_name for wks in self.worksheets]

    def wks(self, wks_id:str|int) -> Worksheet:
        "Find a worksheet by name or number"
        if isinstance(wks_id, int):
            try:
                return self.worksheets[wks_id]
            except IndexError:
                raise KeyError(f"Cannot find worksheet {wks_id}")
        else:
            for worksheet in self.worksheets:
                if worksheet.sheet_name == wks_id:
                    return worksheet
            raise KeyError(f"Cannot find worksheet '{wks_id}'")
        


    @property
    def header_colors(self):
        "Return the header colors for each tab"
        return {wks.sheet_name: wks.header_color for wks in self.worksheets}

    @property
    def emphasizes(self):
        "Return the emphasis for each tab"
        return {wks.sheet_name: wks.emphasize for wks in self.worksheets}

    @property
    def tab_colors(self):
        "Return the colors for each tab"
        return {wks.sheet_name: wks.tab_color for wks in self.worksheets}


   
    def xl_number_formats(self, wks:Worksheet) -> dict:
        """
        Utility method.

        Apply default formats + specified ones to a worksheet.
        This has to be done here, since the conventions are defined
        at report level.
        """
        formats = {}
        for col, spec in wks.columns.items():
            # print(col, spec)
            if spec == "int":
                formats[col] = self.format_int
            elif spec == 'float':
                formats[col] = self.format_float
            elif spec == 'perc':
                formats[col] = self.format_perc
            elif spec == 'date':
                formats[col] = self.format_date
        # update with the user specs:
        formats.update(wks.num_formats)
        return formats


    @property
    def number_formats(self) -> dict:
        """
        Return all the number formats for each tab
        We apply default formats to columns.
        """
        return {wks.sheet_name: self.xl_number_formats(wks)
                for wks in self.worksheets}


    @property
    def images(self):
        "Return the emphasis for each tab"
        d = {}
        for wks in self.worksheets:
            sheet_name = wks.sheet_name
            if wks.image_filename:
                d[sheet_name] = SheetImage(wks.image_filename,
                                           wks.image_position)
            else:
                d[sheet_name] = None
        return d
        

    def save(self, open_file:bool=False) -> openpyxl.Workbook:
        """
        Saves the whole workbook onto a file.
        Returns the openpyxl workbook (in case)
        """
        # First, save with panda's format
        assert self.filename is not None
        with  pd.ExcelWriter(self.filename) as writer:
            for wks in self.worksheets:
               wks.df.to_excel(writer, 
                                sheet_name=wks.sheet_name, 
                                index=False)
        # print("Header color:", self.header_color)
        wb = format_xl_report(filename=self.filename, 
                        header_colors=self.header_colors,
                        font_name = self.font_name,
                        number_formats=self.number_formats,
                        emphasizes=self.emphasizes,
                        tab_colors=self.tab_colors,
                        images=self.images,
                        debug=self.debug)
        # assign
        self.workbook = wb
        if open_file:
            self.open()
        return wb

    def open(self):
        "Open the Excel file"
        open_file(self.filename)

    def rich_print(self, wks_id=0, printing:bool=True):
        "Rich print the first table on the console"
        try:
            wks = self.wks(wks_id)
            return format_text_table(wks.df,
                    header_color=wks.header_color or self.header_color,
                    printing=printing)
        except IndexError as e:
            raise IndexError("No worksheet available on this report.")
    
    def __str__(self):
        "Printing the first table"
        return self.rich_print(printing=False)

