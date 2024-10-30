"Utilities for Excel reports"

import os, sys, subprocess
import re
from functools import lru_cache
from datetime import datetime


# import xlsxwriter

import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import filedialog
from webcolors import name_to_hex

# Column width
MAX_COLUMN_SIZE = 25 # Maximum no of characters in a row


# --------------------------------
# Create Excel workbook (table)
# --------------------------------

def get_columns(header, rows):
    """
    Get a list columns from header (list of cells) 
    and rows (list of lists of cells)
    """
    table = [header] + rows
    r = list(zip(*table))
    assert len(header) == len(r)
    try:
        assert len(table) == len(r[2])
    except IndexError:
        pass
    return r

def get_col_widths(header, rows):
    "Find the maximum length of each column (heuristically)"
    EXCEL_WIDTH_FACTOR = 1.2 # factor for calculating column width
    r = []
    # add some margin on headers (for filter arrows):
    header2 = [str(text) + '___' for text in header]
    # Go through each column:
    cols = get_columns(header2, rows)
    for col in cols:
        #assert len(col) > 2
        chars = max(len(str(s)) + 1 for s in col)
        # limit size:
        if chars > MAX_COLUMN_SIZE:
            chars = MAX_COLUMN_SIZE
        r.append(chars * EXCEL_WIDTH_FACTOR)
        # print(col, "=>", chars)
    return r

def xl_col_to_name(col_num:int) -> str:
    """
    Convert a zero-indexed column number to an Excel column name.
    """
    name = ''
    while col_num >= 0:
        name = chr(col_num % 26 + ord('A')) + name
        col_num = col_num // 26 - 1
    return name


# --------------------------------
# Workon colors
# --------------------------------
@lru_cache
def to_rgb(color:str) -> tuple[int]:
    "Get the rgb"
    if not color or not isinstance(color, str):
        raise TypeError(f"Argument is not a valid string")
    if color[0] != '#':
        hex_color = name_to_hex(color)
    else:
        hex_color = color
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

@lru_cache
def to_argb(color:str):
    """
    Convert name or hexa to aRGB
    https://www.delftstack.com/howto/python/python-hex-to-rgb/
    """
    r = '%02x%02x%02x' % to_rgb(color)
    return r.upper()



def get_luminance(r, g, b):
    "Get luminance from RGB"
    return 0.299 * r + 0.587 * g + 0.114 * b



def get_font_color(background_color:str) -> str:
    """
    Decide black or white color according to color indicated
    (name or hex)
    """
    r, g, b = to_rgb(background_color)
    luminance = get_luminance(r, g, b)
    return 'black' if luminance > 128 else 'white'

# --------------------------------
# Format Excel workbook
# This is low level
# --------------------------------
# for openpyxl (ex-post formatting of tables produced by pandas)
def get_column_id(column_names, col:str, return_int:bool=False):
    """
    Get the column id (e.g. 'B3') from the column name or index (int)

    Returns
    -------
    By default a string (e.g. 'A', 'AA').
    
    If return_int is True, returns an int, in Excel format
    (starting from 1)
    """
    if isinstance(col, int):
        col_id = col
    else:
        # get the index number
        col_id = column_names.index(col)
    if return_int:
        return col_id + 1 # Excel
    else:
        # default case: return a column letter (A, B...)
        # return xlsxwriter.utility.xl_col_to_name(col_id)
        return get_column_letter(col_id + 1)
        
def get_wks(workbook: openpyxl.workbook, tab=0):
    """ 
    Get a worksheet from a workbook
    
    Arguments
    ---------
    - workbook: the workbook (object)
    - tab: name or index of the worksheet 

    Returns
    -------
    worksheet object
    """
    if isinstance(workbook, int):
        sheets = workbook.sheetnames
        tab = sheets[tab]
    return workbook[tab]

def expand_name(name:str, actual:list[str], desc='item',
                regex_marker:str='//',
                case_sensitive:bool=False) -> list[str]:
    """
    Expand a regex name to match those
    contained in another (actual).
    If no equivalence is found, fail with Value Error.

    For simplicity, the regex is case insensitive by default.

    Arguments
    ---------
    name: the name to be matched (will be treated as regex if starts with '//')
        if it is an int (index), it will converted into its str value, 
        if possible
    actual: the reference list
    desc: descriptive name to be used in the error message
        ('filename', 'col', 'row', etc.)
    regex_marker: the prefix for a regex if different from '//'
    case_sensitive: if the regex is case sensitive.

    Returns
    -------
    List of items

    Examples:
    ---------
    expand_name('//foo', ['foo', 'bar', 'baz', 'foobar', 'No Fool]) -> 
        ['foo', 'foobar', 'No Fool']
    """
    # make sure its all string, or fail
    if not all([isinstance(el, str) for el in actual]):
        raise TypeError(f"Not all {desc}s are strings! ({actual})")
    if isinstance(name, int):
        # it's an index -- if empty forward the int
        return [actual[name]] or name
    if  name.startswith(regex_marker):
        found = []
        exp = name[len(regex_marker):] # remove prefix
        flags = 0
        if not case_sensitive:
            flags = flags | re.IGNORECASE
        parser = re.compile(exp, flags) # regexp
        # print("EXPAND:", name, actual)
        for actual_el in actual:
            # sometimes col names can be ints!
            if parser.search(actual_el):
                    # print("  Found:", actual_el)
                    found.append(actual_el)
        if not found:
            raise ValueError(f"No {desc} matching '{name}' was found")
        return found
    else:
        # normal name
        if not name in actual:
            raise ValueError(f"No {desc} '{name}' was found {actual}")
        else:
            return [name]
        


# --------------------------------
# Date mapping
# --------------------------------
FORMAT_DATETIME_MAP = {
        'YYYY-MM-DD': '%Y-%m-%d',
        'DD/MM/YYYY': '%d/%m/%Y',
        'MM/DD/YYYY': '%m/%d/%Y',
        'DD-MMM-YYYY': '%d-%b-%Y',
        'MMM DD, YYYY': '%b %d, %Y',
        'MMMM DD, YYYY': '%B %d, %Y',
        'YYYY-MM-DD HH:MM': '%Y-%m-%d %H:%M',
        'DD/MM/YYYY HH:MM': '%d/%m/%Y %H:%M',
        'MM/DD/YYYY HH:MM': '%m/%d/%Y %H:%M',
        'DD-MMM-YYYY HH:MM': '%d-%b-%Y %H:%M',
        'MMM DD, YYYY HH:MM': '%b %d, %Y %H:%M',
        'MMMM DD, YYYY HH:MM': '%B %d, %Y %H:%M',
        'YYYY-MM-DD HH:MM:SS': '%Y-%m-%d %H:%M:%S',
        'DD/MM/YYYY HH:MM:SS': '%d/%m/%Y %H:%M:%S',
        'MM/DD/YYYY HH:MM:SS': '%m/%d/%Y %H:%M:%S',
        'DD-MMM-YYYY HH:MM:SS': '%d-%b-%Y %H:%M:%S',
        'MMM DD, YYYY HH:MM:SS': '%b %d, %Y %H:%M:%S',
        'MMMM DD, YYYY HH:MM:SS': '%B %d, %Y %H:%M:%S',
        'HH:MM': '%H:%M',
        'HH:MM:SS': '%H:%M:%S',
}

def xl_strftime(dt:datetime, style:str)-> str:
    """
    Translate an Excel date format to a Python strftime format.

    Parameters:
    dt: the date to convert
    style: The Excel date format (e.g., 'YYYY-MM-DD').

    Returns:
    str: The corresponding Python strftime format.
    """
    style = style.strip()
    try:
        python_format = FORMAT_DATETIME_MAP[style]
    except KeyError:
        raise KeyboardInterrupt(f"Sorry, I can interpret '{style}'")
    return dt.strftime(python_format)


# --------------------------------
# Other utilities
# --------------------------------

if sys.platform == 'win32':
    import win32.lib.win32con as win32con
    import win32api
    import win32event
    from win32com.shell import shellcon
    from win32com.shell.shell import ShellExecuteEx

    def start_file_wait(fname):
        "Windows: start file and wait"
        rc = ShellExecuteEx(
            fMask=shellcon.SEE_MASK_NOCLOSEPROCESS,
            nShow=win32con.SW_SHOW,
            lpFile=fname)
        hproc = rc['hProcess']
        win32event.WaitForSingleObject(hproc, win32event.INFINITE)
        win32api.CloseHandle(hproc)

def open_file(filename: str, wait=False) -> str:
    """
    Open a file with the usual application

    Arguments
    ---------
    filename: the name of the file to be opened; 
        if directory name is provided, it opens a file browser at that location.
    wait: wait until the application exits, to continue
    """
    if os.path.isdir(filename):
        filename = filedialog.askopenfilename(
            initialdir=filename,
            title="Select a Document")
        if not filename:
            return ''

    if os.path.isfile(filename):
        try:  # should work on Windows
            if wait:
                start_file_wait(filename)
            else:
                os.startfile(filename)
        except AttributeError:
            try:  # should work on MacOS and most linux versions
                if sys.platform == 'darwin':
                    command = ['open', filename]
                elif sys.platform == 'linux':
                    command = ['xdg-open', filename]
                if wait:
                    command.insert(1, '-W')
                subprocess.call(command)
            except:
                raise FileNotFoundError(f"Problem opening '{filename}'")
    else:
        raise FileNotFoundError(f"File '{filename} does not exist'")
    return filename