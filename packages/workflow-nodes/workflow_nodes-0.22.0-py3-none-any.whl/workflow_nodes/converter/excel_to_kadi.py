# Copyright 2020 Karlsruhe Institute of Technology
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
import os
import sys

import click
from kadi_apy import RESOURCE_ROLES
from kadi_apy import apy_command
from kadi_apy.cli.commons import validate_metadatum
from openpyxl import load_workbook
from openpyxl.utils import cell
from openpyxl.utils import get_column_letter
from xmlhelpy import Choice
from xmlhelpy import Path
from xmlhelpy import argument
from xmlhelpy import option

from .main import converter


def _read_value(ws, row, column):
    value = ws.cell(row=row, column=column).value
    if value is None:
        return value
    return str(value).strip()


def _parse_list(value):
    if value is None:
        return value
    value = value.split(";")
    value = map(str.strip, value)
    value = [obj for obj in value if obj]
    return value


def _check_int(value, output):
    try:
        value = int(value)
    except ValueError:
        click.echo(f"No valid integer given in {output}.")
        sys.exit(1)

    return value


@converter.command()
@apy_command
@argument(
    name="file",
    description="The Excel file to read from.",
    param_type=Path(path_type="file", exists=True),
)
@option(
    "force",
    char="f",
    description="Force deleting and overwriting existing information.",
    is_flag=True,
)
@option(
    "keep-formulas",
    char="k",
    description="Flag to indicate whether to print the formula of a given cell, even if"
    " a computed value is available.",
    is_flag=True,
)
@option("start-column", char="S", description="Start column.", default="E")
@option("end-column", char="E", description="End column.")
@option(
    "permission-new",
    char="p",
    description="Permission of new user or group.",
    default="member",
    param_type=Choice(RESOURCE_ROLES["record"]),
)
@option(
    "base-path",
    char="b",
    description="Prefix path to be added in front of the files or paths specified in"
    " the Excel sheet.",
)
@option(
    "use-ids",
    char="U",
    description="Use ids to specify collections, groups and users.",
    is_flag=True,
)
@option(
    "identity-type",
    char="D",
    description="Identity type of the user.",
    param_type=Choice(["ldap", "local", "shib"]),
)
@option(
    "metadatum-with-value",
    char="M",
    description="Consider only metadata which have a value.",
    is_flag=True,
)
def excel_to_kadi(
    manager,
    file,
    force,
    keep_formulas,
    start_column,
    end_column,
    base_path,
    permission_new,
    use_ids,
    identity_type,
    metadatum_with_value,
):
    """Imports an Excel sheet, reads metadata and transfers them into Kadi."""
    wb = load_workbook(filename=file, read_only=True, data_only=not keep_formulas)
    ws = wb.active

    if not end_column:
        end_column = ws.max_column
    else:
        end_column = cell.column_index_from_string(end_column)

    start_column_int = cell.column_index_from_string(start_column)

    for i in range(start_column_int, end_column + 1):
        identifier = _read_value(ws, 1, i)
        if identifier is None:
            break

        title = _read_value(ws, 2, i)
        description = _read_value(ws, 3, i)
        type = _read_value(ws, 4, i)
        tags = _parse_list(_read_value(ws, 5, i))
        add_collections = _parse_list(_read_value(ws, 6, i))
        add_groups = _parse_list(_read_value(ws, 7, i))
        add_user = _parse_list(_read_value(ws, 8, i))
        links = _parse_list(_read_value(ws, 10, i))
        title_links = _parse_list(_read_value(ws, 11, i))
        files = _parse_list(_read_value(ws, 13, i))

        if base_path:
            if base_path[-1] != os.sep:
                base_path = base_path + os.sep

        record = manager.record(identifier=identifier, title=title, create=True)

        if description:
            record.set_attribute(description=description)

        if type:
            record.set_attribute(type=type)

        if tags:
            for tag in tags:
                record.add_tag(tag)

        if add_collections:
            for collection_iter in add_collections:
                if use_ids:
                    _check_int(collection_iter, f"{get_column_letter(i)}6")
                    collection = manager.collection(id=collection_iter)
                else:
                    collection = manager.collection(identifier=collection_iter)
                record.add_collection_link(collection=collection)

        if add_groups:
            for group_iter in add_groups:
                if use_ids:
                    _check_int(group_iter, f"{get_column_letter(i)}7")
                    group = manager.group(id=group_iter)
                else:
                    group = manager.group(identifier=group_iter)
                record.add_group_role(group=group, permission_new=permission_new)

        if add_user:
            for user_iter in add_user:
                if use_ids:
                    _check_int(user_iter, f"{get_column_letter(i)}8")
                    user = manager.user(id=user_iter)
                else:
                    if identity_type is None:
                        click.echo(
                            "Please specify the identity type via the option '-D'."
                        )
                        sys.exit(1)
                    user = manager.user(username=user_iter, identity_type=identity_type)
                record.add_user(user=user, permission_new=permission_new)

        if files:
            for obj in files:
                if base_path:
                    obj = base_path + obj
                record.upload_file(file_name=obj, force=force, pattern="*")

        if links:
            if len(links) != len(title_links):
                click.echo(
                    f"Found {len(links)} entries for links but {len(title_links)}"
                    " titles. Please use the same number of entries."
                )
                sys.exit(1)
            for link_iter, title_iter in zip(links, title_links):
                if use_ids:
                    _check_int(link_iter, f"{get_column_letter(i)}10")
                    record_to = manager.record(id=link_iter)
                else:
                    record_to = manager.record(identifier=link_iter)
                record.link_record(record_to=record_to, name=title_iter)

        metadata = []
        x = 17
        while True:
            metadatum_key = _read_value(ws, x, cell.column_index_from_string("A"))
            if not metadatum_key:
                break
            metadatum_value = _read_value(ws, x, i)
            metadatum_type = _read_value(ws, x, cell.column_index_from_string("C"))
            metadatum_unit = _read_value(ws, x, cell.column_index_from_string("D"))

            x = x + 1

            if metadatum_value is None and metadatum_with_value:
                continue

            metadata.append(
                validate_metadatum(
                    metadatum=metadatum_key,
                    value=metadatum_value,
                    type=metadatum_type,
                    unit=metadatum_unit,
                )
            )

        record.add_metadata(metadata=metadata, force=force)
